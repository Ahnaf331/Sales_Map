import pandas as pd
import numpy as np
import folium
from folium.plugins import MarkerCluster, HeatMap, MiniMap
import random
import json
import argparse
import os
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ======================
# Config Loader
# ======================
def load_config(config_file="config.json"):
    with open(config_file, "r", encoding="utf-8") as f:
        return json.load(f)


# ======================
# Synthetic Data Generator
# ======================
def generate_synthetic_sales(config, n_records=1000, seed=42):
    random.seed(seed)
    np.random.seed(seed)

    categories = config["categories"]
    price_ranges = config["price_ranges"]
    cities = config["cities"]
    category_probs = config["category_probs"]
    mean_units = config["mean_units"]

    # Normalize city weights
    city_weights = np.array([c[3] for c in cities], dtype=float)
    city_weights /= city_weights.sum()

    city_names = [c[0] for c in cities]
    city_lats = np.array([c[1] for c in cities])
    city_lons = np.array([c[2] for c in cities])

    cat_names = list(category_probs.keys())
    cat_probs = np.array([category_probs[k] for k in cat_names])

    rows = []
    today = date.today()

    for outlet_id in range(1, n_records + 1):
        # Pick city
        idx = np.random.choice(len(city_names), p=city_weights)
        city, lat0, lon0 = city_names[idx], city_lats[idx], city_lons[idx]

        latitude = float(lat0 + np.random.normal(scale=0.06))
        longitude = float(lon0 + np.random.normal(scale=0.08))

        # Category & product
        category = str(np.random.choice(cat_names, p=cat_probs))
        product = str(random.choice(categories[category]))

        # Price
        low, high = price_ranges[category]
        base_price = float(np.random.uniform(low, high))
        price = int(np.round(base_price * np.random.uniform(0.85, 1.25)))

        # Units sold (seasonality: boost groceries during Ramadan months)
        multiplier = float(np.clip(np.random.normal(1.0, 0.35), 0.3, 2.5))
        lam = max(0.5, mean_units[category] * multiplier)
        units_sold = int(np.random.poisson(lam))

        # Sales
        sales_amount = int(units_sold * price)

        # Sale date (within last year)
        days_ago = int(np.random.randint(0, 365))
        sale_date = (today - timedelta(days=days_ago)).isoformat()

        # Store attributes
        store_size = str(np.random.choice(['Small', 'Medium', 'Large'], p=[0.5, 0.35, 0.15]))
        customer_rating = float(np.round(np.clip(np.random.normal(4.1, 0.5), 1.0, 5.0), 2))
        returns = int(np.random.binomial(units_sold, 0.02)) if units_sold > 0 else 0

        rows.append({
            'outlet_id': outlet_id,
            'city': city,
            'latitude': latitude,
            'longitude': longitude,
            'region': city,
            'category': category,
            'product': product,
            'price_bdt': price,
            'units_sold': units_sold,
            'sales_amount': sales_amount,
            'sale_date': sale_date,
            'store_size': store_size,
            'customer_rating': customer_rating,
            'returns': returns
        })

    df = pd.DataFrame(rows)
    df['sales_per_unit'] = df.apply(lambda r: round(r['sales_amount'] / r['units_sold'], 2) if r['units_sold'] > 0 else 0.0, axis=1)

    # Ensure outputs directory exists
    os.makedirs('outputs', exist_ok=True)
    
    csv_path = 'outputs/synthetic_sales_data.csv'
    df.to_csv(csv_path, index=False)
    print(f"✅ Saved synthetic dataset: {csv_path} ({len(df)} rows)")
    return df


# ======================
# Map Creator
# ======================
def create_interactive_map(df, config, out_html='outputs/sales_map.html'):
    CENTER = [23.7, 90.4]
    m = folium.Map(location=CENTER, zoom_start=6, tiles='CartoDB dark_matter')

    category_color_map = config["category_colors"]

    markers = MarkerCluster(name='Outlets (clustered)').add_to(m)
    max_sales = df['sales_amount'].max() if df['sales_amount'].max() > 0 else 1

    for _, r in df.iterrows():
        popup_html = (
            f"<div style='font-size:12px'>"
            f"<b>Outlet:</b> {r['outlet_id']}<br/>"
            f"<b>City:</b> {r['city']}<br/>"
            f"<b>Category:</b> {r['category']}<br/>"
            f"<b>Product:</b> {r['product']}<br/>"
            f"<b>Units sold:</b> {r['units_sold']}<br/>"
            f"<b>Sales (BDT):</b> {int(r['sales_amount']):,}<br/>"
            f"<b>Date:</b> {r['sale_date']}<br/>"
            f"<b>Rating:</b> {r['customer_rating']}⭐<br/>"
            f"<b>Returns:</b> {r['returns']}"
            f"</div>"
        )

        # Marker scaling (log scale helps outliers)
        radius = float(3 + (np.log1p(r['sales_amount']) / np.log1p(max_sales)) * 12)
        radius = max(3, min(16, radius))

        folium.CircleMarker(
            location=(r['latitude'], r['longitude']),
            radius=radius,
            popup=folium.Popup(popup_html, max_width=320),
            tooltip=f"{r['product']} — {r['category']} — {int(r['sales_amount']):,} BDT",
            color=None,
            fill=True,
            fill_color=category_color_map.get(r['category'], '#333333'),
            fill_opacity=0.75,
            weight=0.5
        ).add_to(markers)

    # Heatmap
    heat_in = [[row['latitude'], row['longitude'], float(row['sales_amount'] / max_sales)] for _, row in df.iterrows()]
    heat_group = folium.FeatureGroup(name='Sales Heatmap')
    HeatMap(heat_in, radius=18, blur=12, max_zoom=7).add_to(heat_group)
    m.add_child(heat_group)

    # MiniMap + Layer Control with dark theme
    minimap = MiniMap(toggle_display=True, tile_layer='CartoDB dark_matter')
    m.add_child(minimap)
    folium.LayerControl(collapsed=False).add_to(m)

    # Legend with dark theme
    legend_html = """<div style="position: fixed; 
                                bottom: 30px; left: 30px; width: 180px; 
                                border:2px solid #555; z-index:9999; font-size:12px;
                                background-color:rgba(40, 40, 40, 0.95); 
                                color: white;
                                padding: 15px;
                                border-radius: 8px;
                                box-shadow: 0 4px 6px rgba(0,0,0,0.5);">
                        <b style="font-size: 14px;">Category Colors</b><br><br>"""
    for cat, color in category_color_map.items():
        legend_html += f'<i style="background:{color};width:12px;height:12px;float:left;margin-right:8px;border-radius:2px;"></i>{cat}<br>'
    legend_html += "</div>"
    m.get_root().html.add_child(folium.Element(legend_html))

    m.save(out_html)
    print(f"✅ Saved interactive map: {out_html}")
    return df  # Return the dataframe for Excel export


# ======================
# Excel Export
# ======================
def export_to_excel(df, filename='outputs/sales_analysis.xlsx'):
    """Export all data and summaries to an Excel file with multiple sheets."""
    print("🔹 Generating Excel report...")
    
    # Create a Pandas Excel writer using openpyxl
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    
    # Main data sheet
    df['sale_date'] = pd.to_datetime(df['sale_date'])
    df['month'] = df['sale_date'].dt.strftime('%Y-%m')
    
    # Reorder columns for better readability
    columns_order = [
        'outlet_id', 'city', 'region', 'store_size', 'customer_rating',
        'category', 'product', 'price_bdt', 'units_sold', 'sales_amount',
        'sales_per_unit', 'returns', 'sale_date', 'month', 'latitude', 'longitude'
    ]
    df = df[columns_order].sort_values(['city', 'category', 'product'])
    
    # Write main data
    df.to_excel(writer, sheet_name='Sales Data', index=False)
    
    # Create summary sheets
    create_summary_sheets(writer, df)
    
    # Get workbook and worksheet objects
    workbook = writer.book
    worksheet = writer.sheets['Sales Data']
    
    # Format the main data sheet
    format_worksheet(worksheet, df.columns)
    
    # Auto-adjust column widths
    for i, col in enumerate(df.columns, 1):
        max_length = max(df[col].astype(str).apply(len).max(), len(str(col))) + 2
        worksheet.column_dimensions[get_column_letter(i)].width = min(max_length, 25)
    
    # Save the Excel file
    writer.close()
    print(f"✅ Excel report saved: {filename}")


def create_summary_sheets(writer, df):
    """Create summary sheets in the Excel file."""
    # Summary by City
    city_summary = df.groupby(['city', 'region']).agg({
        'outlet_id': 'count',
        'sales_amount': 'sum',
        'units_sold': 'sum',
        'returns': 'sum'
    }).reset_index()
    city_summary = city_summary.rename(columns={'outlet_id': 'outlet_count'})
    city_summary['avg_sale_per_outlet'] = city_summary['sales_amount'] / city_summary['outlet_count']
    city_summary['return_rate'] = (city_summary['returns'] / city_summary['units_sold'] * 100).round(2)
    city_summary.to_excel(writer, sheet_name='City Summary', index=False)
    
    # Summary by Category
    category_summary = df.groupby('category').agg({
        'outlet_id': 'count',
        'sales_amount': 'sum',
        'units_sold': 'sum',
        'price_bdt': 'mean'
    }).reset_index()
    category_summary = category_summary.rename(columns={
        'outlet_id': 'transaction_count',
        'price_bdt': 'avg_price'
    })
    category_summary.to_excel(writer, sheet_name='Category Summary', index=False)
    
    # Monthly Summary
    monthly_summary = df.groupby('month').agg({
        'outlet_id': 'count',
        'sales_amount': 'sum',
        'units_sold': 'sum'
    }).reset_index()
    monthly_summary = monthly_summary.rename(columns={'outlet_id': 'transaction_count'})
    monthly_summary['avg_sale_per_transaction'] = monthly_summary['sales_amount'] / monthly_summary['transaction_count']
    monthly_summary.to_excel(writer, sheet_name='Monthly Summary', index=False)
    
    # Get workbook and format summary sheets
    workbook = writer.book
    for sheet_name in ['City Summary', 'Category Summary', 'Monthly Summary']:
        worksheet = writer.sheets[sheet_name]
        header_row = [cell.value for cell in worksheet[1]]
        format_worksheet(worksheet, header_row)
        
        # Auto-adjust column widths
        for col_idx in range(1, len(header_row) + 1):
            max_length = 0
            # Check header width first
            header_cell = worksheet.cell(row=1, column=col_idx)
            max_length = len(str(header_cell.value)) if header_cell.value else 0
            
            # Check data rows
            for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_value))
            
            # Set column width with some padding
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 30)


def format_worksheet(worksheet, columns):
    """Apply formatting to a worksheet."""
    # Format header
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'
    
    # Add filter
    worksheet.auto_filter.ref = worksheet.dimensions


# ======================
# Summaries
# ======================
def summarize_and_export(df):
    # This function is kept for backward compatibility
    # All functionality moved to export_to_excel
    return None, None, None


# ======================
# Main
# ======================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--records", type=int, default=1000, help="Number of synthetic records to generate")
    parser.add_argument("--config", type=str, default="config.json", help="Path to config file")
    args = parser.parse_args()

    config = load_config(args.config)

    print("🔹 Generating synthetic dataset...")
    df = generate_synthetic_sales(config, n_records=args.records)
    
    print("🔹 Creating interactive Folium map...")
    df = create_interactive_map(df, config)  # Get the updated dataframe
    
    print("🔹 Exporting data to Excel...")
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_file = f'outputs/sales_analysis_{timestamp}.xlsx'
    export_to_excel(df, excel_file)
    
    print("\n🎉 All done. Outputs in 'outputs/' folder.")
    print(f"   - Interactive map: outputs/sales_map.html")
    print(f"   - Excel report: {excel_file}")
    print(f"   - Raw data: outputs/synthetic_sales_data.csv")


if __name__ == '__main__':
    main()
